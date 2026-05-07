"""Excel import/export helpers for Dizimistas."""
from datetime import datetime, timezone
from io import BytesIO
import uuid

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def _header_styles():
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="b91c1c", end_color="b91c1c", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    return header_font, header_fill, header_align, thin_border


def build_template_workbook() -> BytesIO:
    """Generate a downloadable Excel template for dizimistas import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dizimistas"

    header_font, header_fill, header_align, thin_border = _header_styles()
    headers = [
        "Nome*", "Telefone Celular", "Telefone Residencial", "Email",
        "Logradouro", "Número", "Complemento", "CEP",
        "Data Nascimento (DD/MM/AAAA)", "Valor Dízimo"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    example = [
        "João da Silva", "(11) 99999-9999", "(11) 2222-3333", "joao@email.com",
        "Rua das Flores", "123", "Apto 45", "03456-000", "15/06/1980", "100.00"
    ]
    for col, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.font = Font(italic=True, color="888888")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


COLUMN_ALIASES = {
    "nome": ["nome", "name", "nome*"],
    "telefone": ["celular", "telefone celular", "telefone", "tel", "tel celular", "phone"],
    "telefone_residencial": ["tel. residencial", "telefone residencial", "tel residencial", "residencial"],
    "email": ["email", "e-mail", "mail"],
    "logradouro": ["logradouro", "endereço", "endereco", "rua", "address"],
    "numero": ["nº", "número", "numero", "num", "n"],
    "complemento": ["complemento", "compl", "apto", "apartamento"],
    "cep": ["cep", "zip", "código postal", "codigo postal"],
    "data_nascimento": ["aniv", "aniversário", "aniversario", "data nascimento", "data nascimento (dd/mm/aaaa)", "nascimento", "birthday"],
    "nota": ["nota", "observação", "observacao", "note"],
    "status": ["status", "situação", "situacao"],
    "comunicacao": ["comunicação", "comunicacao", "contato preferido"],
    "valor_dizimo": ["valor dízimo", "valor dizimo", "valor", "dízimo", "dizimo"],
    "estado_civil": ["estado civil", "civil"],
    "conjuge": ["cônjuge", "conjuge", "spouse"],
    "co_dizimista": ["co-dizimista", "co dizimista", "codizimista", "co_dizimista"],
    "co_dizimista_aniversario": [
        "aniv. co-dizimista", "aniversário co-dizimista", "aniv co-dizimista",
        "aniversario co-dizimista", "co_dizimista_aniversario"
    ],
}


def _parse_date(raw):
    if not raw:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    date_str = str(raw).strip()
    parts = date_str.replace("-", "/").split("/")
    if len(parts) == 3:
        return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    if len(parts) == 2:
        current_year = datetime.now().year
        return f"{current_year}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return ""


def _parse_valor(raw):
    if not raw:
        return 0.0
    try:
        return float(str(raw).replace(",", ".").replace("R$", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _validate_choice(value, valid_options, default):
    value = str(value or "").strip()
    return value if value in valid_options else default


def parse_workbook(content: bytes):
    """Parse an uploaded Excel file and return list of dizimista dicts + errors."""
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

    col_map = {}
    for field, aliases in COLUMN_ALIASES.items():
        for i, header in enumerate(headers):
            if header in aliases:
                col_map[field] = i
                break

    def get_cell(row, field, default=""):
        if field in col_map and col_map[field] < len(row):
            val = row[col_map[field]]
            return val if val is not None else default
        return default

    dizimistas = []
    errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        nome = get_cell(row, "nome", "")
        if not nome:
            continue
        try:
            dizimista = {
                "id": str(uuid.uuid4()),
                "nome": str(nome).strip(),
                "telefone": str(get_cell(row, "telefone", "")).strip(),
                "telefone_residencial": str(get_cell(row, "telefone_residencial", "")).strip(),
                "email": str(get_cell(row, "email", "")).strip(),
                "logradouro": str(get_cell(row, "logradouro", "")).strip(),
                "numero": str(get_cell(row, "numero", "")).strip(),
                "complemento": str(get_cell(row, "complemento", "")).strip(),
                "cep": str(get_cell(row, "cep", "")).strip(),
                "data_nascimento": _parse_date(get_cell(row, "data_nascimento", "")),
                "co_dizimista": str(get_cell(row, "co_dizimista", "")).strip(),
                "co_dizimista_aniversario": _parse_date(get_cell(row, "co_dizimista_aniversario", "")),
                "nota": _validate_choice(get_cell(row, "nota", ""), ["Novo", "Atualizar", "OK"], "Novo"),
                "status": _validate_choice(get_cell(row, "status", ""), ["Ativo", "Pendente", "Inativo"], "Ativo"),
                "comunicacao": _validate_choice(get_cell(row, "comunicacao", ""), ["WhatsApp", "Correio", "E-mail", ""], ""),
                "valor_dizimo": _parse_valor(get_cell(row, "valor_dizimo", "")),
                "data_cadastro": datetime.now(timezone.utc).isoformat(),
                "ultima_contribuicao": "",
            }
            dizimistas.append(dizimista)
        except Exception as e:
            errors.append(f"Linha {row_num}: {str(e)}")
    return dizimistas, errors


def build_export_workbook(dizimistas: list) -> BytesIO:
    """Build a styled Excel workbook with the given dizimistas list."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Dizimistas"

    header_font, header_fill, header_align, thin_border = _header_styles()
    headers = [
        "Nome", "Co-Dizimista", "Aniv. Co-Dizimista", "Celular",
        "Tel. Residencial", "Email", "Logradouro", "Nº", "Complemento", "CEP",
        "Aniversário", "Nota", "Status", "Estado Civil", "Cônjuge", "Comunicação"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    column_widths = {
        'A': 30, 'B': 25, 'C': 15, 'D': 15, 'E': 15, 'F': 25,
        'G': 30, 'H': 8, 'I': 15, 'J': 12, 'K': 12, 'L': 10,
        'M': 10, 'N': 12, 'O': 25, 'P': 12,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    def _format_dm(date_str):
        if not date_str:
            return ""
        try:
            parts = date_str.split("-")
            if len(parts) == 3:
                return f"{parts[2]}/{parts[1]}"
        except Exception:
            pass
        return ""

    for row_num, d in enumerate(dizimistas, 2):
        values = [
            d.get("nome", ""),
            d.get("co_dizimista", ""),
            _format_dm(d.get("co_dizimista_aniversario", "")),
            d.get("telefone", ""),
            d.get("telefone_residencial", ""),
            d.get("email", ""),
            d.get("logradouro", d.get("endereco", "")),
            d.get("numero", ""),
            d.get("complemento", ""),
            d.get("cep", ""),
            _format_dm(d.get("data_nascimento", "")),
            d.get("nota", ""),
            d.get("status", "Ativo"),
            d.get("estado_civil", ""),
            d.get("nome_conjuge", ""),
            d.get("comunicacao", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
