from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
SECRET_KEY = os.environ.get('JWT_SECRET', 'igreja-dizimistas-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class UserPermissions(BaseModel):
    dizimistas_view: bool = False
    dizimistas_edit: bool = False
    relatorios_view: bool = False
    relatorios_edit: bool = False

class UserBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    name: str
    role: str = "user"  # admin or user
    permissions: UserPermissions = Field(default_factory=UserPermissions)
    active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str = "user"
    permissions: Optional[UserPermissions] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    permissions: Optional[UserPermissions] = None
    active: Optional[bool] = None

class UserLogin(BaseModel):
    username: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

class DizimistaBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    telefone: str = ""
    telefone_residencial: str = ""
    email: str = ""
    logradouro: str = ""
    numero: str = ""
    complemento: str = ""
    cep: str = ""
    data_nascimento: str = ""
    nota: str = "Novo"  # Atualizar, Novo, OK
    status: str = "Ativo"  # Ativo, Pendente, Inativo
    modo_contribuicao: str = ""  # PIX, Envelope, Depósito
    mes_contribuicao: str = ""  # Mês preferencial de contribuição
    comunicacao: str = ""  # WhatsApp, Correio, E-mail
    valor_dizimo: float = 0.0
    data_cadastro: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    ultima_contribuicao: str = ""

class DizimistaCreate(BaseModel):
    nome: str
    telefone: str = ""
    telefone_residencial: str = ""
    email: str = ""
    logradouro: str = ""
    numero: str = ""
    complemento: str = ""
    cep: str = ""
    data_nascimento: str = ""
    nota: str = "Novo"
    status: str = "Ativo"
    modo_contribuicao: str = ""
    mes_contribuicao: str = ""
    comunicacao: str = ""
    valor_dizimo: float = 0.0

class DizimistaUpdate(BaseModel):
    nome: Optional[str] = None
    telefone: Optional[str] = None
    telefone_residencial: Optional[str] = None
    email: Optional[str] = None
    logradouro: Optional[str] = None
    numero: Optional[str] = None
    complemento: Optional[str] = None
    cep: Optional[str] = None
    data_nascimento: Optional[str] = None
    nota: Optional[str] = None
    status: Optional[str] = None
    modo_contribuicao: Optional[str] = None
    mes_contribuicao: Optional[str] = None
    comunicacao: Optional[str] = None
    valor_dizimo: Optional[float] = None

class ContribuicaoBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dizimista_id: str
    dizimista_nome: str = ""  # Para facilitar visualização
    valor: float
    data: str = Field(default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    mes_referencia: str = ""  # Mês de referência da contribuição
    meio: str = ""  # Envelope, Pix/Depósito, Presencial

class ContribuicaoCreate(BaseModel):
    dizimista_id: str
    valor: float
    data: Optional[str] = None
    mes_referencia: str = ""
    meio: str = ""

class ContribuicaoUpdate(BaseModel):
    valor: Optional[float] = None
    data: Optional[str] = None
    mes_referencia: Optional[str] = None
    meio: Optional[str] = None

class ValorMensalBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    mes: int  # 1-12
    ano: int
    valor: float
    observacao: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ValorMensalCreate(BaseModel):
    mes: int
    ano: int
    valor: float
    observacao: str = ""

class ValorMensalUpdate(BaseModel):
    valor: Optional[float] = None
    observacao: Optional[str] = None

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")

def check_permission(user: dict, resource: str, action: str) -> bool:
    if user.get("role") == "admin":
        return True
    permissions = user.get("permissions", {})
    key = f"{resource}_{action}"
    return permissions.get(key, False)

# Initialize admin user
@app.on_event("startup")
async def startup_db():
    admin = await db.users.find_one({"username": "admin"})
    if not admin:
        admin_user = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password": hash_password("admin123"),
            "name": "Administrador",
            "role": "admin",
            "permissions": {
                "dizimistas_view": True,
                "dizimistas_edit": True,
                "relatorios_view": True,
                "relatorios_edit": True
            },
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        logging.info("Admin user created: admin/admin123")

# Auth Routes
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"username": data.username}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    if not user.get("active", True):
        raise HTTPException(status_code=401, detail="Usuário desativado")
    
    token = create_access_token({"sub": user["id"]})
    user_data = {k: v for k, v in user.items() if k != "password"}
    return TokenResponse(access_token=token, user=user_data)

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# User Management Routes (Admin only)
@api_router.get("/users")
async def list_users(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.post("/users")
async def create_user(data: UserCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Usuário já existe")
    
    permissions = data.permissions.model_dump() if data.permissions else {
        "dizimistas_view": False,
        "dizimistas_edit": False,
        "relatorios_view": False,
        "relatorios_edit": False
    }
    
    user = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "password": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "permissions": permissions,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    user.pop("password", None)
    user.pop("_id", None)
    return user

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if "permissions" in update_data and update_data["permissions"]:
        update_data["permissions"] = update_data["permissions"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if user_id == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Não é possível excluir o próprio usuário")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": "Usuário excluído com sucesso"}

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, permissions: UserPermissions, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": permissions.model_dump()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return user

# Dizimistas Routes
# Excel routes first (before dynamic routes)
@api_router.get("/dizimistas/template/excel")
async def download_template(current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dizimistas"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="b91c1c", end_color="b91c1c", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Nome*", "Telefone Celular", "Telefone Residencial", "Email", "Logradouro", "Número", "Complemento", "CEP", "Data Nascimento (DD/MM/AAAA)", "Valor Dízimo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Example row
    example = ["João da Silva", "(11) 99999-9999", "(11) 2222-3333", "joao@email.com", "Rua das Flores", "123", "Apto 45", "03456-000", "15/06/1980", "100.00"]
    for col, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.font = Font(italic=True, color="888888")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_dizimistas.xlsx"}
    )

@api_router.post("/dizimistas/import/excel")
async def import_dizimistas_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para importar dizimistas")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")
    
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    imported = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:  # Skip empty rows
            continue
        
        try:
            # Parse date (column 8 now)
            data_nasc = ""
            if len(row) > 8 and row[8]:
                if isinstance(row[8], datetime):
                    data_nasc = row[8].strftime("%Y-%m-%d")
                else:
                    parts = str(row[8]).split("/")
                    if len(parts) == 3:
                        data_nasc = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            
            # Parse valor (column 9 now)
            valor = 0.0
            if len(row) > 9 and row[9]:
                valor = float(str(row[9]).replace(",", "."))
            
            dizimista = {
                "id": str(uuid.uuid4()),
                "nome": str(row[0]).strip(),
                "telefone": str(row[1] or "").strip(),
                "telefone_residencial": str(row[2] or "").strip(),
                "email": str(row[3] or "").strip(),
                "logradouro": str(row[4] or "").strip(),
                "numero": str(row[5] or "").strip(),
                "complemento": str(row[6] or "").strip(),
                "cep": str(row[7] or "").strip(),
                "data_nascimento": data_nasc,
                "nota": "Novo",
                "status": "Ativo",
                "valor_dizimo": valor,
                "data_cadastro": datetime.now(timezone.utc).isoformat(),
                "ultima_contribuicao": ""
            }
            await db.dizimistas.insert_one(dizimista)
            imported += 1
        except Exception as e:
            errors.append(f"Linha {row_num}: {str(e)}")
    
    return {"imported": imported, "errors": errors}

@api_router.get("/dizimistas/export/excel")
async def export_dizimistas_excel(
    status: Optional[str] = None,
    nota: Optional[str] = None,
    mes_aniversario: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    if not check_permission(current_user, "dizimistas", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    query = {}
    if status and status != "todos":
        query["status"] = status
    if nota and nota != "todos":
        query["nota"] = nota
    
    dizimistas = await db.dizimistas.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by birthday month
    if mes_aniversario:
        filtered = []
        for d in dizimistas:
            if d.get("data_nascimento"):
                try:
                    parts = d["data_nascimento"].split("-")
                    if len(parts) >= 2 and int(parts[1]) == mes_aniversario:
                        filtered.append(d)
                except:
                    pass
        dizimistas = filtered
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Dizimistas"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="b91c1c", end_color="b91c1c", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Nome", "Celular", "Tel. Residencial", "Email", "Logradouro", "Nº", "Complemento", "CEP", "Aniversário", "Nota", "Status", "Modo Contrib.", "Comunicação", "Valor Dízimo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    
    for row_num, d in enumerate(dizimistas, 2):
        # Format birthday
        aniversario = ""
        if d.get("data_nascimento"):
            try:
                parts = d["data_nascimento"].split("-")
                if len(parts) == 3:
                    aniversario = f"{parts[2]}/{parts[1]}"
            except:
                pass
        
        values = [
            d.get("nome", ""),
            d.get("telefone", ""),
            d.get("telefone_residencial", ""),
            d.get("email", ""),
            d.get("logradouro", d.get("endereco", "")),
            d.get("numero", ""),
            d.get("complemento", ""),
            d.get("cep", ""),
            aniversario,
            d.get("nota", ""),
            d.get("status", "Ativo"),
            d.get("modo_contribuicao", ""),
            d.get("comunicacao", ""),
            d.get("valor_dizimo", 0)
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "lista_dizimistas"
    if status and status != "todos":
        filename += f"_{status}"
    if nota and nota != "todos":
        filename += f"_{nota}"
    if mes_aniversario:
        meses = ["", "janeiro", "fevereiro", "marco", "abril", "maio", "junho", 
                 "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
        filename += f"_aniversario_{meses[mes_aniversario]}"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )

@api_router.get("/dizimistas")
async def list_dizimistas(
    nota: Optional[str] = None,
    status: Optional[str] = None,
    mes_aniversario: Optional[int] = None,
    nome: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not check_permission(current_user, "dizimistas", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar dizimistas")
    
    query = {}
    if nota and nota != "todos":
        query["nota"] = nota
    if status and status != "todos":
        query["status"] = status
    if nome:
        query["nome"] = {"$regex": nome, "$options": "i"}
    
    dizimistas = await db.dizimistas.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by birthday month and sort by day
    if mes_aniversario:
        filtered = []
        for d in dizimistas:
            if d.get("data_nascimento"):
                try:
                    parts = d["data_nascimento"].split("-")
                    if len(parts) >= 2 and int(parts[1]) == mes_aniversario:
                        # Add day for sorting
                        d["_sort_day"] = int(parts[2]) if len(parts) >= 3 else 0
                        filtered.append(d)
                except:
                    pass
        # Sort by day (menor para maior)
        filtered.sort(key=lambda x: x.get("_sort_day", 0))
        # Remove temporary sort field
        for d in filtered:
            d.pop("_sort_day", None)
        dizimistas = filtered
    
    return dizimistas

@api_router.post("/dizimistas")
async def create_dizimista(data: DizimistaCreate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para criar dizimistas")
    
    dizimista = DizimistaBase(**data.model_dump()).model_dump()
    await db.dizimistas.insert_one(dizimista)
    dizimista.pop("_id", None)
    return dizimista

@api_router.put("/dizimistas/{dizimista_id}")
async def update_dizimista(dizimista_id: str, data: DizimistaUpdate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para editar dizimistas")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    
    result = await db.dizimistas.update_one({"id": dizimista_id}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Dizimista não encontrado")
    
    dizimista = await db.dizimistas.find_one({"id": dizimista_id}, {"_id": 0})
    return dizimista

@api_router.delete("/dizimistas/{dizimista_id}")
async def delete_dizimista(dizimista_id: str, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para excluir dizimistas")
    
    result = await db.dizimistas.delete_one({"id": dizimista_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dizimista não encontrado")
    return {"message": "Dizimista excluído com sucesso"}

# Contribuicoes Routes
@api_router.get("/contribuicoes")
async def list_contribuicoes(
    dizimista_id: Optional[str] = None,
    mes_referencia: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not check_permission(current_user, "dizimistas", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar contribuições")
    
    query = {}
    if dizimista_id:
        query["dizimista_id"] = dizimista_id
    if mes_referencia and mes_referencia != "todos":
        query["mes_referencia"] = mes_referencia
    
    contribuicoes = await db.contribuicoes.find(query, {"_id": 0}).sort("data", -1).to_list(10000)
    
    # Filter by date range
    if data_inicio:
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:10] >= data_inicio]
    if data_fim:
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:10] <= data_fim]
    
    # Add dizimista name to each contribution
    dizimistas = await db.dizimistas.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(10000)
    diz_map = {d["id"]: d["nome"] for d in dizimistas}
    
    for c in contribuicoes:
        c["dizimista_nome"] = diz_map.get(c.get("dizimista_id"), "Desconhecido")
    
    return contribuicoes

@api_router.post("/contribuicoes")
async def create_contribuicao(data: ContribuicaoCreate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para registrar contribuições")
    
    # Get dizimista name
    dizimista = await db.dizimistas.find_one({"id": data.dizimista_id}, {"_id": 0})
    if not dizimista:
        raise HTTPException(status_code=404, detail="Dizimista não encontrado")
    
    contribuicao_data = data.model_dump()
    contribuicao_data["dizimista_nome"] = dizimista.get("nome", "")
    
    if not contribuicao_data.get("data"):
        contribuicao_data["data"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    contribuicao = ContribuicaoBase(**contribuicao_data).model_dump()
    await db.contribuicoes.insert_one(contribuicao)
    contribuicao.pop("_id", None)
    
    # Update dizimista status to Ativo and ultima_contribuicao
    await db.dizimistas.update_one(
        {"id": data.dizimista_id},
        {"$set": {
            "status": "Ativo",
            "ultima_contribuicao": contribuicao_data["data"]
        }}
    )
    
    return contribuicao

@api_router.put("/contribuicoes/{contribuicao_id}")
async def update_contribuicao(contribuicao_id: str, data: ContribuicaoUpdate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para editar contribuições")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    
    result = await db.contribuicoes.update_one({"id": contribuicao_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contribuição não encontrada")
    
    updated = await db.contribuicoes.find_one({"id": contribuicao_id}, {"_id": 0})
    return updated

@api_router.delete("/contribuicoes/{contribuicao_id}")
async def delete_contribuicao(contribuicao_id: str, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "dizimistas", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para excluir contribuições")
    
    result = await db.contribuicoes.delete_one({"id": contribuicao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contribuição não encontrada")
    
    return {"message": "Contribuição excluída"}

# Update dizimistas status based on contributions
async def update_all_dizimistas_status():
    """Update status of all dizimistas based on their last contribution date"""
    now = datetime.now(timezone.utc)
    dizimistas = await db.dizimistas.find({}, {"_id": 0}).to_list(10000)
    
    for d in dizimistas:
        ultima = d.get("ultima_contribuicao", "")
        if not ultima:
            # Check contributions collection for this dizimista
            last_contrib = await db.contribuicoes.find_one(
                {"dizimista_id": d["id"]},
                sort=[("data", -1)]
            )
            if last_contrib:
                ultima = last_contrib.get("data", "")
        
        if ultima:
            try:
                last_date = datetime.fromisoformat(ultima.replace("Z", "+00:00"))
                months_diff = (now.year - last_date.year) * 12 + (now.month - last_date.month)
                
                if months_diff >= 6:
                    new_status = "Inativo"
                elif months_diff >= 2:
                    new_status = "Pendente"
                else:
                    new_status = "Ativo"
                
                if d.get("status") != new_status:
                    await db.dizimistas.update_one(
                        {"id": d["id"]},
                        {"$set": {"status": new_status, "ultima_contribuicao": ultima}}
                    )
            except:
                pass

@api_router.post("/dizimistas/atualizar-status")
async def trigger_status_update(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores podem atualizar status")
    
    await update_all_dizimistas_status()
    return {"message": "Status dos dizimistas atualizado com sucesso"}

# Relatorios Routes
@api_router.get("/relatorios/resumo")
async def get_resumo(
    mes_inicio: Optional[str] = None,
    ano_inicio: Optional[str] = None,
    mes_fim: Optional[str] = None,
    ano_fim: Optional[str] = None,
    status: Optional[str] = None,
    nota: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not check_permission(current_user, "relatorios", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar relatórios")
    
    # Get dizimistas with optional filters
    diz_query = {}
    if status and status != "todos":
        diz_query["status"] = status
    if nota and nota != "todos":
        diz_query["nota"] = nota
    
    dizimistas = await db.dizimistas.find(diz_query, {"_id": 0}).to_list(10000)
    dizimistas_ids = set(d["id"] for d in dizimistas)
    total_dizimistas = len(dizimistas)
    
    # Count by status (always from all dizimistas, not filtered)
    all_dizimistas = await db.dizimistas.find({}, {"_id": 0}).to_list(10000)
    dizimistas_ativos = len([d for d in all_dizimistas if d.get("status") == "Ativo"])
    dizimistas_pendentes = len([d for d in all_dizimistas if d.get("status") == "Pendente"])
    dizimistas_inativos = len([d for d in all_dizimistas if d.get("status") == "Inativo"])
    total_geral = len(all_dizimistas)
    
    # Calculate total from dizimistas valor_dizimo
    total_valor_dizimo = sum(d.get("valor_dizimo", 0) for d in all_dizimistas if d.get("status") == "Ativo")
    
    # Get contributions
    contribuicoes = await db.contribuicoes.find({}, {"_id": 0}).to_list(10000)
    valores_mensais = await db.valores_mensais.find({}, {"_id": 0}).to_list(10000)
    
    # Filter contributions by dizimista (if status/nota filters applied)
    if status or nota:
        contribuicoes = [c for c in contribuicoes if c.get("dizimista_id") in dizimistas_ids]
    
    # Filter by period
    if ano_inicio and mes_inicio:
        data_inicio = f"{ano_inicio}-{mes_inicio.zfill(2)}"
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:7] >= data_inicio]
        valores_mensais = [v for v in valores_mensais if f"{v.get('ano')}-{str(v.get('mes')).zfill(2)}" >= data_inicio]
    
    if ano_fim and mes_fim:
        data_fim = f"{ano_fim}-{mes_fim.zfill(2)}"
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:7] <= data_fim]
        valores_mensais = [v for v in valores_mensais if f"{v.get('ano')}-{str(v.get('mes')).zfill(2)}" <= data_fim]
    
    total_contribuicoes = sum(c.get("valor", 0) for c in contribuicoes)
    total_valores_mensais = sum(v.get("valor", 0) for v in valores_mensais)
    
    # Total arrecadado = soma dos valores de dízimo + contribuições registradas + valores mensais
    total_arrecadado = total_valor_dizimo + total_contribuicoes + total_valores_mensais
    
    # Monthly breakdown
    monthly = {}
    for c in contribuicoes:
        data = c.get("data", "")[:7]
        if data:
            monthly[data] = monthly.get(data, 0) + c.get("valor", 0)
    
    for v in valores_mensais:
        key = f"{v.get('ano')}-{str(v.get('mes')).zfill(2)}"
        monthly[key] = monthly.get(key, 0) + v.get("valor", 0)
    
    return {
        "total_dizimistas": total_geral,
        "dizimistas_ativos": dizimistas_ativos,
        "dizimistas_pendentes": dizimistas_pendentes,
        "dizimistas_inativos": dizimistas_inativos,
        "total_arrecadado": total_arrecadado,
        "total_valor_dizimo": total_valor_dizimo,
        "total_contribuicoes": len(contribuicoes),
        "por_mes": monthly
    }

@api_router.get("/relatorios/contribuicoes")
async def get_relatorio_contribuicoes(
    mes_inicio: Optional[str] = None,
    ano_inicio: Optional[str] = None,
    mes_fim: Optional[str] = None,
    ano_fim: Optional[str] = None,
    status: Optional[str] = None,
    nota: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not check_permission(current_user, "relatorios", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar relatórios")
    
    # Get all contributions
    contribuicoes = await db.contribuicoes.find({}, {"_id": 0}).to_list(10000)
    
    # Get dizimistas with status and nota
    dizimistas = await db.dizimistas.find({}, {"_id": 0}).to_list(10000)
    dizimistas_map = {d["id"]: d for d in dizimistas}
    
    # Filter by period
    if ano_inicio and mes_inicio:
        data_inicio = f"{ano_inicio}-{mes_inicio.zfill(2)}"
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:7] >= data_inicio]
    
    if ano_fim and mes_fim:
        data_fim = f"{ano_fim}-{mes_fim.zfill(2)}"
        contribuicoes = [c for c in contribuicoes if c.get("data", "")[:7] <= data_fim]
    
    # Filter by status and nota (from dizimista)
    if status or nota:
        filtered = []
        for c in contribuicoes:
            diz = dizimistas_map.get(c.get("dizimista_id"), {})
            if status and diz.get("status") != status:
                continue
            if nota and diz.get("nota") != nota:
                continue
            filtered.append(c)
        contribuicoes = filtered
    
    # Add dizimista info to contributions
    for c in contribuicoes:
        diz = dizimistas_map.get(c.get("dizimista_id"), {})
        c["dizimista_nome"] = diz.get("nome", "Desconhecido")
        c["dizimista_status"] = diz.get("status", "")
        c["dizimista_nota"] = diz.get("nota", "")
    
    return contribuicoes

# Valores Mensais Routes (Historical monthly totals)
@api_router.get("/valores-mensais")
async def list_valores_mensais(current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "relatorios", "view"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar valores mensais")
    valores = await db.valores_mensais.find({}, {"_id": 0}).to_list(1000)
    return valores

@api_router.post("/valores-mensais")
async def create_valor_mensal(data: ValorMensalCreate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "relatorios", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para registrar valores mensais")
    
    # Check if already exists for this month/year
    existing = await db.valores_mensais.find_one({"mes": data.mes, "ano": data.ano})
    if existing:
        # Update existing
        await db.valores_mensais.update_one(
            {"mes": data.mes, "ano": data.ano},
            {"$set": {"valor": data.valor, "observacao": data.observacao}}
        )
        updated = await db.valores_mensais.find_one({"mes": data.mes, "ano": data.ano}, {"_id": 0})
        return updated
    
    valor = ValorMensalBase(**data.model_dump()).model_dump()
    await db.valores_mensais.insert_one(valor)
    valor.pop("_id", None)
    return valor

@api_router.put("/valores-mensais/{valor_id}")
async def update_valor_mensal(valor_id: str, data: ValorMensalUpdate, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "relatorios", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para editar valores mensais")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    
    result = await db.valores_mensais.update_one({"id": valor_id}, {"$set": update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Valor mensal não encontrado")
    
    valor = await db.valores_mensais.find_one({"id": valor_id}, {"_id": 0})
    return valor

@api_router.delete("/valores-mensais/{valor_id}")
async def delete_valor_mensal(valor_id: str, current_user: dict = Depends(get_current_user)):
    if not check_permission(current_user, "relatorios", "edit"):
        raise HTTPException(status_code=403, detail="Sem permissão para excluir valores mensais")
    
    result = await db.valores_mensais.delete_one({"id": valor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Valor mensal não encontrado")
    return {"message": "Valor mensal excluído com sucesso"}

# Health check
@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
